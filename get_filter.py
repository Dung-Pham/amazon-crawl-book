from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import time
import pandas as pd
from openpyxl import load_workbook



# Thiết lập tùy chọn trình duyệt với user-agent
options = webdriver.ChromeOptions()
options.add_argument('--headless')  # Tùy chọn chạy không giao diện
options.add_argument('--disable-gpu')
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_argument(
    'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3')

# Khởi tạo ChromeDriver với Service và ChromeOptions
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)

base_url = 'https://www.amazon.com/amz-books/discover?node=1&navStore=books&bbn=1000&rh=n%3A283155%2Cn%3A1&dc&qid=1726471931&rnid=1000&ref=lp_1000_nr_n_0&ref_=lp_1000_nr_n_0'

# page_url = f"{base_url}&page={page}"
def getListTopic(file_path, columns):
    
    # Đọc file Excel
    # file_path = 'path_to_your_file.xlsx'
    df = pd.read_excel(file_path, sheet_name='Sheet1', usecols = columns)  # Chỉnh sửa sheet_name nếu cần
    # columns = ['id', 'link']  # Thay đổi tên cột theo nhu cầu của bạn

    # Chuyển DataFrame thành danh sách
    data_list = df.to_dict(orient='records')  # Mỗi phần tử trong danh sách là một từ điển

    return data_list
    
    
def getFilters(base_url, Tid) :
    driver.get(base_url)
    try: 
        filter_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'div.css-x0ph25'))
        )
        filter_button.click()

        

        filter_shadow_host = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'tag-filters-section[data-testid="default_genre-modal-filters"]'))
        )
        filter_shadow_root =  driver.execute_script('return arguments[0].shadowRoot', filter_shadow_host)
        elFilter = filter_shadow_root.find_element(By.CSS_SELECTOR, 'div.tagFiltersSection')

        filter_html = elFilter.get_attribute("outerHTML")
        soup = BeautifulSoup(filter_html, 'lxml')
        #lấy danh sách các filter
        filters = soup.find_all('li')
        i = 1
        for fil in filters:
            tag_filter = fil.find('tag-pill')
            filter = tag_filter.get('tagname')
            
            filter_info = {
                'Fid': Tid + "F" + str(i),
                'Filter': filter
            }
            i+=1
            filter_data.append(filter_info)
        # df = pd.DataFrame(filter_data)
        # df.to_excel('filters_data.xlsx', index=False, engine='openpyxl')
    except :
        print('Trường hợp không có List Filters...')
        print('----------------------------------------------------------------------------------------------------')

    return filter_data


def updateData(new_data, file_path):
    df = pd.DataFrame(new_data)

    # file_path = 'topic_data1.xlsx'
    
    # Tải workbook và chọn sheet
    workbook = load_workbook(file_path)
    sheet = workbook['Sheet1']

    # Xác định hàng tiếp theo để thêm dữ liệu
    next_row = sheet.max_row + 1

    # Thêm nhiều hàng vào sheet cùng một lúc
    for index, row in df.iterrows():
        sheet.append(row.tolist())

    # Lưu workbook
    workbook.save(file_path)

filter_data = []
# Đọc danh sách topic từ file topic_data.xlsx
columns = ['Tid', 'link']
file_path = 'data/topic_data.xlsx'
list_topic = getListTopic(file_path, columns)

for topic in list_topic:
# # topic = list_topic[1]
    print(topic)
    new_data = getFilters( topic["link"], topic["Tid"])
    updateData(new_data, 'filters_data.xlsx')
    print('Đã lấy thành công filter của topic ' + topic["Tid"] )
    print('Tổng số filter thu được: ',len(new_data))
    print('----------------------------------------------------------------------------------------------------')
    new_data.clear()


# finally:
#     driver.quit()