from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import time
import pandas as pd
from openpyxl import load_workbook


# Thiết lập tùy chọn trình duyệt với user-agent
options = webdriver.ChromeOptions()
options.add_argument('--headless')  # Tùy chọn chạy không giao diện
options.add_argument('--disable-gpu')
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_argument(
    'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3')

# Khởi tạo ChromeDriver với Service và ChromeOptions
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)

def getListData(file_path, columns):
    
    # Đọc file Excel
    # file_path = 'path_to_your_file.xlsx'
    df = pd.read_excel(file_path, sheet_name='Sheet1', usecols = columns)  

    # Chuyển DataFrame thành danh sách
    data_list = df.to_dict(orient='records')  # Mỗi phần tử trong danh sách là một từ điển

    return data_list


# Hàm lấy dữ liệu từ 1 trang
# def scrape_page(page_url, books_data):

# page_url = f"{base_url}&page={page}"
def getBooks(base_url, filter_name, fid ):
        
    driver.get(base_url)
    # html = driver.page_source
    # soup = BeautifulSoup(html, 'lxml')

    # with open('test1.html', 'w', encoding='utf-8') as file:
    #     file.write(soup.prettify())

    books_data = []

    try:
        try:
            filter_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'div.css-x0ph25'))
            )
            filter_button.click()
            
            filter_shadow_host = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'tag-filters-section[data-testid="default_genre-modal-filters"]'))
            )
            filter_shadow_root =  driver.execute_script('return arguments[0].shadowRoot', filter_shadow_host)
        
            # Sử dụng JavaScript để truy cập shadow-root
            # shadow_root = driver.execute_script('return arguments[0].shadowRoot', shadow_host)
            # Tìm phần tử bên trong shadow DOM
            filter_list = filter_shadow_root.find_element(By.CSS_SELECTOR, 'div.tagFiltersSection')
            
            filter_button = filter_list.find_element(By.CSS_SELECTOR, f'tag-pill[tagname="{filter_name}"]')
            
            driver.execute_script("arguments[0].scrollIntoView(true);", filter_button)
            filter_button.click()   
            # driver.execute_script("arguments[0].click();", filter_button)
        # except:
        #     print('Không thể mở filter...')
        except Exception as e:
            print(f"Đã xảy ra lỗi huhu: {e}")
        show_results = driver.find_element(By.CSS_SELECTOR, 'button.css-1p5aykz')
        show_results.click()

        while True:
            try:
                WebDriverWait(driver, 10).until(
                    # EC.presence_of_element_located((By.CSS_SELECTOR, "span.a-size-medium.a-color-base.a-text-normal")))
                    EC.presence_of_element_located((By.XPATH, "//button[@aria-label='Back to top of the tag bar']")))
                print('Đã mở toàn bộ sách!')
                break
            except Exception as e:
                print("Đang mở rộng...")
                
            button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Show more recommendations']")))
            # print('đã tìm thấy nút show more!')
            button.click()
            
        # Lấy nội dung trang và phân tích với BeautifulSoup
        # html = driver.page_source
        # time.sleep(5)

        html = driver.execute_script("return document.documentElement.outerHTML;")
        soup = BeautifulSoup(html, 'lxml')

        with open('test.html', 'w', encoding='utf-8') as file:
            file.write(soup.prettify())
            
        #    Tìm tất cả các sản phẩm
        products = soup.find_all('div', {'class': 'css-1nlt46i'})
        i = 0
        for product in products:
            # Lấy tiêu đề
            imgs = product.find_all('img')
            title = imgs[1].get('alt') if imgs[1] else "Không có tiêu đề"
            
            # Lấy giá bán
            price_tag = product.find('span', {'class': 'css-15zmhrp'})
            price = price_tag.get_text(strip=True) if price_tag else "Không có giá"

            # Lấy đánh giá
            rating_tag = product.find('span', {'class': 'css-1vapab1'})
            rating = rating_tag.get_text(strip=True) if rating_tag else "Không có đánh giá"

            # Lấy số lượt đánh giá
            review_count_tag = product.find('span', {'class': 'css-1bdv1w0'})
            review_count = review_count_tag.get_text(strip=True) if review_count_tag else "Không có số lượt đánh giá"
            
            # Lấy link ảnh
            image = imgs[1].get('srcset').split(",\n")[0] if imgs[1] else "Không có ảnh"
        
            # Lấy đường link
            link_tag = product.find('a')
            link = 'https://www.amazon.com' + link_tag.get('href') if link_tag else "Không có đường link"
            i+= 1
            # Lưu thông tin vào mảng (danh sách) books_data
            book_info = {
                'Bid': fid + "B" + str(i),
                'title': title,
                'price': price,
                'rating': rating,
                # 'author': author,
                'review_count': review_count,
                'link': link,
                'image': image,
                
            }
            books_data.append(book_info)
            
        # In ra dữ liệu đã cào theo dạng [sản phẩm, tiêu đề, giá, đánh giá]
        # for idx, book in enumerate(books_data, start=1):
        #     print([{idx}, book['title']])

        # df = pd.DataFrame(books_data)
        # df.to_excel('data/books_data.xlsx', index=False, engine='openpyxl')


    except Exception as e:
        print(f"Đã xảy ra lỗi: {e}")
    return books_data

def updateData(new_data, file_path):
    df = pd.DataFrame(new_data)

    # file_path = 'topic_data1.xlsx'
    
    # Tải workbook và chọn sheet
    workbook = load_workbook(file_path)
    sheet = workbook['Sheet1']

    # Xác định hàng tiếp theo để thêm dữ liệu
    next_row = sheet.max_row + 1

    # Thêm nhiều hàng vào sheet cùng một lúc
    for index, row in df.iterrows():
        sheet.append(row.tolist())

    # Lưu workbook
    workbook.save(file_path)
    

Book_data = []

# Đọc danh sách topic từ file topic_data.xlsx
topic_columns = ['Tid', 'link']
filter_columns = ['Fid', 'Filter']

topic_file_path = 'data/topic_data.xlsx'
filter_file_path = 'data/filters_data.xlsx'
# danh sách topic
list_topic = getListData(topic_file_path, topic_columns)
# danh sách filter
list_filter = getListData(filter_file_path, filter_columns)

# base_url = 'https://www.amazon.com/amz-books/discover?node=1&navStore=books&bbn=1000&rh=n%3A283155%2Cn%3A1&dc&qid=1726471931&rnid=1000&ref=lp_1000_nr_n_0&ref_=lp_1000_nr_n_0'

for filter in list_filter[:40]: # vì vừa chạy từ 0, vừa bị cắt
    
    Fid = filter['Fid']
    Tid = int(Fid[1:Fid.find('F')])
    print('Topic: ', Tid)
    print('Filter: ', Fid)

    topic = list_topic[(Tid - 1)]
# print(topic)
    new_data = getBooks(topic['link'], filter['Filter'], Fid)
    updateData(new_data, 'data/books_data.xlsx')
    print('Đã lấy thành công Book của filter ' + filter["Fid"] )
    print('Tổng số Book thu được: ',len(new_data))
    print('----------------------------------------------------------------------------------------------------')
    new_data.clear()


        
# finally:
#    driver.quit()